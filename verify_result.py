#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""결과 확인"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_133710.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 80)
print("첫 번째 데이터 행 (행 2) 결과 확인")
print("=" * 80)

print(f"\n입력 데이터:")
print(f"  Project (D2): {ws['D2'].value}")
print(f"  CVE (L2): {ws['L2'].value}")
print(f"  GHSA (M2): {ws['M2'].value}")
print(f"  ANT (N2): {ws['N2'].value}")

print(f"\n자동 채워진 데이터:")
print(f"  R 열 (취약 버전): {ws['R2'].value}")
print(f"  S 열 (조치 버전): {ws['S2'].value}")

print("\n" + "=" * 80)
print("몇 가지 추가 행 샘플")
print("=" * 80)

for row in [3, 7, 8, 15, 20]:
    print(f"\n[행 {row}]")
    print(f"  L: {ws[f'L{row}'].value}")
    print(f"  R: {ws[f'R{row}'].value}")
    print(f"  S: {ws[f'S{row}'].value}")

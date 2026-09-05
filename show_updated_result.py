#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""최종 결과 확인 - unaffected 버전 포함"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_141959.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 140)
print("최종 결과 - R 열과 S 열 (S 열에 unaffected 버전 포함)")
print("=" * 140)

sample_rows = [2, 7, 8, 9, 15, 20]

for row in sample_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    print(f"\n[행 {row}] CVE: {cve}")

    if r_col:
        r_display = str(r_col)[:100] + "..." if len(str(r_col)) > 100 else str(r_col)
        print(f"  R 열 (취약 버전):")
        print(f"    {r_display}")
    else:
        print(f"  R 열 (취약 버전): (없음)")

    if s_col:
        s_display = str(s_col)[:100] + "..." if len(str(s_col)) > 100 else str(s_col)
        print(f"  S 열 (조치/unaffected 버전):")
        print(f"    {s_display}")
    else:
        print(f"  S 열 (조치/unaffected 버전): (없음)")

print("\n" + "=" * 140)
print("✓ 완료. S 열에 unaffected 버전이 포함되었습니다.")
print("=" * 140)

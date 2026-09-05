#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Nexpose 자료정리.xlsx — R/S 열 자동 채우기 (고도화 버전)
프로젝트명으로 GitHub 보안 공지 검색 + NVD API 조회
Affected Products 테이블만 사용 (통일된 데이터 소스)
"""

import openpyxl
import requests
import json
import time
from pathlib import Path
import logging
from typing import Tuple, Optional
from openpyxl.styles import Alignment

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ 설정 ============
EXCEL_FILE = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리.xlsx"
CVE_COL = "L"
GHSA_COL = "M"
ANT_COL = "N"
PROJECT_COL = "D"
R_COL = "R"       # 취약 버전
S_COL = "S"       # 조치 버전

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
GITHUB_API = "https://api.github.com"
GITHUB_TOKEN = None  # os.getenv("GITHUB_TOKEN")  # 선택사항, 없어도 동작

# ============ API 함수 ============

def get_nvd_info(cve_id: str) -> Tuple[Optional[str], Optional[str]]:
    """NVD API로 CVE 정보 조회 - Affected Products 테이블만 사용"""
    try:
        url = f"{NVD_API}?cveId={cve_id}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        if not data.get('vulnerabilities'):
            return None, None

        vuln = data['vulnerabilities'][0]['cve']
        affected_list = []
        fixed_list = []

        # Affected Products 테이블 (affected 배열)에서만 추출
        for affected_item in vuln.get('affected', []):
            for affected_data in affected_item.get('affectedData', []):
                product = affected_data.get('product', '')

                for version_info in affected_data.get('versions', []):
                    v = version_info.get('version', '')
                    status = version_info.get('status', '')
                    less_equal = version_info.get('lessThanOrEqual', '')
                    less_than = version_info.get('lessThan', '')
                    ver_type = version_info.get('versionType', '')

                    # R열: Affected 버전 범위
                    if status == 'affected' and (less_equal or less_than):
                        end_version = less_equal or less_than
                        # 와일드카드 제외
                        if end_version and '*' not in end_version:
                            version_range = f"< {end_version}" if less_than else f"<= {end_version}"
                            if v != '0':
                                version_range = f"{v} ~ {end_version}"
                            if product:
                                version_range = f"{product} {version_range}"
                            affected_list.append(version_range)

                    # S열: Unaffected 버전 (패치된 버전)
                    if status == 'unaffected' and v and v != '0':
                        # git commit은 명시적으로 표시
                        if 'git' in ver_type:
                            fixed_list.append(f"{product} git: {v}" if product else f"git: {v}")
                        else:
                            fixed_list.append(f"{product} {v}" if product else v)

        affected_str = "\n".join(sorted(set(affected_list))) if affected_list else None
        fixed_str = "\n".join(fixed_list) if fixed_list else None

        if affected_str:
            logger.info(f"✓ {cve_id}: {affected_str[:50]}...")
        else:
            logger.warning(f"⚠ {cve_id}: 버전 정보 파싱 실패")

        return affected_str, fixed_str if affected_str else None

    except Exception as e:
        logger.error(f"✗ NVD {cve_id}: {e}")
        return None, None


def get_github_advisory(ghsa_id: str) -> Tuple[Optional[str], Optional[str]]:
    """GitHub GHSA 정보 조회"""
    try:
        url = f"{GITHUB_API}/advisories/{ghsa_id}"
        headers = {"Accept": "application/vnd.github.v3+json"}
        if GITHUB_TOKEN:
            headers["Authorization"] = f"token {GITHUB_TOKEN}"

        resp = requests.get(url, headers=headers, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        affected = data.get('affected_range', 'N/A')
        fixed = data.get('fixed_in', 'N/A')

        logger.info(f"✓ {ghsa_id}: Affected={affected}, Fixed={fixed}")
        return affected, fixed

    except Exception as e:
        logger.error(f"✗ GitHub {ghsa_id}: {e}")
        return None, None


def search_github_advisories(project: str) -> Tuple[Optional[str], Optional[str]]:
    """
    프로젝트명으로 GitHub 보안 어드바이저리 검색
    예: "libssh2/libssh2" → GHSA/버전 정보
    """
    try:
        # project = "owner/repo" 형식 가정
        if not project or "/" not in project:
            return None, None

        owner, repo = project.split("/", 1)
        url = f"{GITHUB_API}/repos/{owner}/{repo}/security/advisories"
        headers = {"Accept": "application/vnd.github.v3+json"}
        if GITHUB_TOKEN:
            headers["Authorization"] = f"token {GITHUB_TOKEN}"

        resp = requests.get(url, headers=headers, timeout=10)
        resp.raise_for_status()
        advisories = resp.json()

        if not advisories:
            logger.debug(f"  → {project}: 어드바이저리 없음")
            return None, None

        # 최근 어드바이저리 우선
        for adv in advisories[:3]:
            summary = adv.get('summary', '')
            affected = adv.get('affected_range', 'N/A')
            fixed = adv.get('fixed_in', 'N/A')

            logger.info(f"✓ {project}: {summary[:50]}... | {affected} → {fixed}")
            return affected, fixed

    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            logger.debug(f"  → {project}: 리포지토리 없음")
        else:
            logger.error(f"✗ GitHub {project}: {e}")
    except Exception as e:
        logger.error(f"✗ GitHub {project}: {e}")

    return None, None


# ============ ANT-ID 매핑 ============

ANT_DB = {
    # 글래스윙 보고서에서 수동 추출 필요
    # 포맷: "ANT-ID": ("취약 버전 범위", "패치 버전")
    "ANT-2026-Q5A1RHS0": ("libssh2 <= 1.11.0", "1.11.1+"),
    # 아래에 추가...
}


# ============ 메인 처리 ============

def process_row(ws, row_idx: int) -> None:
    """한 행 처리"""
    cve = ws[f"{CVE_COL}{row_idx}"].value
    ghsa = ws[f"{GHSA_COL}{row_idx}"].value
    ant = ws[f"{ANT_COL}{row_idx}"].value
    project = ws[f"{PROJECT_COL}{row_idx}"].value

    affected, fixed = None, None
    source = None

    # 우선순위: CVE > GHSA > GitHub 검색 > ANT-ID
    if cve and isinstance(cve, str) and cve.startswith("CVE-"):
        logger.info(f"[행 {row_idx}] CVE={cve}")
        affected, fixed = get_nvd_info(cve)
        source = "NVD"

    elif ghsa and isinstance(ghsa, str) and ghsa.startswith("GHSA-"):
        logger.info(f"[행 {row_idx}] GHSA={ghsa}")
        affected, fixed = get_github_advisory(ghsa)
        source = "GitHub Advisory"

    elif project and isinstance(project, str):
        logger.info(f"[행 {row_idx}] GitHub 검색: {project}")
        affected, fixed = search_github_advisories(project)
        source = "GitHub 검색"

    elif ant and isinstance(ant, str) and ant.startswith("ANT-"):
        logger.info(f"[행 {row_idx}] ANT-ID={ant}")
        if ant in ANT_DB:
            affected, fixed = ANT_DB[ant]
            source = "로컬 DB"
        else:
            logger.warning(f"  ANT-ID 미등록: {ant}")
            source = None

    # 결과 기입 (실제 데이터가 있을 때만)
    if affected is not None and source:
        r_cell = ws[f"{R_COL}{row_idx}"]
        s_cell = ws[f"{S_COL}{row_idx}"]

        r_cell.value = affected
        s_cell.value = fixed or "[확인 필요]"

        # 셀의 텍스트 줄 바꿈 활성화
        r_cell.alignment = Alignment(wrap_text=True, vertical='top')
        s_cell.alignment = Alignment(wrap_text=True, vertical='top')

        logger.info(f"  → [{source}] R={affected[:40]}... | S={fixed}")
    elif source:
        logger.warning(f"  [{source}] 버전 정보 획득 실패 - 수동 확인 필요")

    time.sleep(0.5)  # Rate limit 회피


def main():
    logger.info(f"파일 로드: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    logger.info(f"시트: {ws.title}, 행: {ws.max_row}\n")

    # 행 2부터 처리
    for row_idx in range(2, ws.max_row + 1):
        try:
            process_row(ws, row_idx)
        except KeyboardInterrupt:
            logger.info("사용자 중단")
            break
        except Exception as e:
            logger.error(f"행 {row_idx} 오류: {e}")

    # 저장
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = EXCEL_FILE.replace(".xlsx", f"_updated_{timestamp}.xlsx")
    wb.save(output)
    logger.info(f"\n✓ 완료. 저장: {output}")


if __name__ == "__main__":
    main()

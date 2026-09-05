#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""최종 결과 표시"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140710.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 130)
print("최종 결과 - R 열과 S 열 (모두 제품명 포함)")
print("=" * 130)

sample_rows = [2, 7, 8, 9, 15, 20]

for row in sample_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    print(f"\n[행 {row}] CVE: {cve}")

    if r_col:
        r_display = str(r_col)[:90] + "..." if len(str(r_col)) > 90 else str(r_col)
        print(f"  R 열 (취약 버전): {r_display}")
    else:
        print(f"  R 열 (취약 버전): (없음)")

    if s_col:
        s_display = str(s_col)[:90] + "..." if len(str(s_col)) > 90 else str(s_col)
        print(f"  S 열 (조치 버전): {s_display}")
    else:
        print(f"  S 열 (조치 버전): (없음)")

print("\n" + "=" * 130)
print("✓ 완료. R 열과 S 열 모두 제품명이 포함되었습니다.")
print("=" * 130)

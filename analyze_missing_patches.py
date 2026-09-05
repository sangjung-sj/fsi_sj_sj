#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""[공식 공지 참조]인 경우 분석"""

import openpyxl
import requests
import json

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_142231.xlsx"

def get_nvd_data(cve_id: str):
    """NVD에서 데이터 조회"""
    try:
        url = f"{NVD_API}?cveId={cve_id}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except:
        return None

def analyze_missing_patch(data, cve_id) -> dict:
    """패치 버전 정보가 없는 이유 분석"""
    result = {
        'cve': cve_id,
        'product': None,
        'affected_versions': [],
        'unaffected_versions': [],
        'reason': None,
        'references': []
    }

    if not data or not data.get('vulnerabilities'):
        return result

    vuln = data['vulnerabilities'][0]['cve']

    # affected 배열에서 버전 정보 수집
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            if not result['product']:
                result['product'] = affected_data.get('product', 'Unknown')

            for version_info in affected_data.get('versions', []):
                status = version_info.get('status', '')
                v = version_info.get('version', '')
                ver_type = version_info.get('versionType', '')

                if status == 'affected':
                    result['affected_versions'].append(v)
                elif status == 'unaffected':
                    result['unaffected_versions'].append({
                        'version': v,
                        'type': ver_type
                    })

    # 패치 버전 정보가 없는 이유 분석
    if not result['unaffected_versions']:
        result['reason'] = '1. unaffected 버전이 NVD에 기록되지 않음'
    else:
        # unaffected 버전이 있지만 모두 git/timestamp인 경우
        git_or_ts = [v for v in result['unaffected_versions'] if 'git' in v['type'] or v['version'].isdigit()]
        valid = [v for v in result['unaffected_versions'] if 'git' not in v['type'] and not v['version'].isdigit()]

        if not valid and git_or_ts:
            result['reason'] = f"2. unaffected가 있지만 모두 특수 형식 ({', '.join([v['type'] for v in git_or_ts])})"
        elif not valid:
            result['reason'] = '3. 모든 unaffected 버전이 필터링됨'

    # References에서 패치 정보 수집
    for ref in vuln.get('references', [])[:3]:
        url = ref.get('url', '')
        tags = ref.get('tags', [])
        if 'Patch' in tags or 'Release' in tags:
            result['references'].append({
                'url': url[:80] + '...' if len(url) > 80 else url,
                'tags': tags
            })

    return result

# 메인 분석
print("=" * 140)
print("[공식 공지 참조]로 표시된 경우 분석")
print("=" * 140)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

missing_patch_cves = []

# S열에 [공식 공지 참조]가 있는 CVE 찾기
for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    s_col = ws[f"S{row}"].value

    if (cve and isinstance(cve, str) and cve.startswith('CVE-') and
        s_col and '[공식 공지 참조]' in str(s_col)):
        missing_patch_cves.append((row, cve))

print(f"\n총 {len(missing_patch_cves)}개의 [공식 공지 참조] CVE 발견\n")

for row, cve in missing_patch_cves[:5]:  # 처음 5개만 분석
    print(f"\n{'─' * 140}")
    print(f"[행 {row}] {cve}")
    print(f"{'─' * 140}")

    nvd_data = get_nvd_data(cve)
    if not nvd_data:
        print("❌ NVD 조회 실패")
        continue

    analysis = analyze_missing_patch(nvd_data, cve)

    print(f"\n제품: {analysis['product']}")
    print(f"\n영향받는 버전 (취약):")
    if analysis['affected_versions']:
        for v in analysis['affected_versions'][:3]:
            print(f"  - {v}")
        if len(analysis['affected_versions']) > 3:
            print(f"  ... 등 {len(analysis['affected_versions']) - 3}개 더")
    else:
        print(f"  (없음)")

    print(f"\nunaffected 버전 (패치):")
    if analysis['unaffected_versions']:
        for v in analysis['unaffected_versions'][:3]:
            print(f"  - {v['version']} (type: {v['type']})")
        if len(analysis['unaffected_versions']) > 3:
            print(f"  ... 등 {len(analysis['unaffected_versions']) - 3}개 더")
    else:
        print(f"  (없음)")

    print(f"\n패치 정보 없는 이유:")
    print(f"  {analysis['reason']}")

    if analysis['references']:
        print(f"\n참고 링크 (공식 공지):")
        for ref in analysis['references']:
            print(f"  - {ref['url']}")
            print(f"    태그: {', '.join(ref['tags'])}")

print("\n\n" + "=" * 140)
print("분석 결과")
print("=" * 140)

print("""
[공식 공지 참조]가 붙는 이유:

1️⃣  NVD에 unaffected (패치) 버전이 명시되지 않은 경우
    → 공식 보안 공지나 release notes 참고 필요

2️⃣  unaffected가 git commit이나 timestamp인 경우
    → 정확한 패치 버전을 파싱 불가
    → 공식 release 정보 필요

3️⃣  References에 공식 보안 공지나 patch 링크만 있는 경우
    → 직접 확인 필요

해결 방법:
✓ References의 Patch/Release 링크를 따라가기
✓ 제품의 공식 보안 공지 확인
✓ GitHub Release/Changelog 확인
""")

print("=" * 140)

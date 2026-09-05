#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""8번째 행 상세 분석"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_142748.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

row = 8
cve = ws[f"L{row}"].value
r_col = ws[f"R{row}"].value
s_col = ws[f"S{row}"].value

print("=" * 150)
print(f"행 {row} 상세 분석")
print("=" * 150)

print(f"\nCVE: {cve}")
print(f"\nR 열 (영향받는 버전/취약 버전):")
print(f"  {r_col}")

print(f"\nS 열 (영향받지 않는 버전/패치 버전):")
print(f"  {s_col}")

print(f"\n" + "=" * 150)
print("논리 검토")
print("=" * 150)

# R, S 값 파싱
r_parts = str(r_col).split("; ") if r_col else []
s_parts = str(s_col).split("; ") if s_col else []

print(f"\nR 열 분석 (취약한 버전들):")
for i, part in enumerate(r_parts):
    print(f"  {i+1}. {part}")

print(f"\nS 열 분석 (안전한/패치된 버전들):")
for i, part in enumerate(s_parts):
    print(f"  {i+1}. {part}")

print(f"\n" + "=" * 150)
print("문제점 검토")
print("=" * 150)

print("""
✓ R열에 있는 취약 버전들:
  - Firefox < 150.0
  - Firefox 140.0 ~ 140.10.0
  - Firefox < 140.10.0
  - Firefox < 115.35.0

✓ S열에 있는 패치 버전들:
  - Firefox 115.35
  - Firefox 140.10
  - Firefox 140.10

논리 검토:
1. "< 115.35.0"이 취약 → 115.35.0 이상이 안전 (115.35 ✓)
2. "140.0 ~ 140.10.0"이 취약 → 140.10.0 이상이 안전 (140.10 ✓)
3. "< 140.10.0"이 취약 → 140.10.0 이상이 안전 (140.10 ✓)
4. "< 150.0"이 취약 → 150.0 이상이 안전 (150은 S열에 없음 ❌)

❓ 이상한 점:
  - R열에 "< 150.0"이 있는데, S열에 150의 안전 버전이 없음?
  - 아니면 다른 논리적 문제?
""")

print("=" * 150)

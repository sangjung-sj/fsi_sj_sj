#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""정확도 검증"""

import openpyxl
import requests
import json
from typing import Dict, List

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140020.xlsx"

def get_nvd_data(cve_id: str):
    """NVD에서 원본 데이터 조회"""
    try:
        url = f"{NVD_API}?cveId={cve_id}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"  ✗ NVD 조회 실패: {e}")
        return None

def extract_nvd_affected(data: Dict) -> Dict:
    """NVD 원본 데이터에서 영향받는 버전 추출"""
    result = {}

    if not data or not data.get('vulnerabilities'):
        return result

    vuln = data['vulnerabilities'][0]['cve']

    # affected 배열에서 추출
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            product = affected_data.get('product', 'Unknown')
            versions = []

            for version_info in affected_data.get('versions', []):
                status = version_info.get('status', '')
                v = version_info.get('version', '')
                less_eq = version_info.get('lessThanOrEqual', '')
                less_than = version_info.get('lessThan', '')

                if status == 'affected' and (less_eq or less_than):
                    end_ver = less_eq or less_than
                    if end_ver and '*' not in end_ver:
                        if v and v != '0':
                            versions.append(f"{v} ~ {end_ver}")
                        else:
                            versions.append(f"<= {end_ver}" if less_eq else f"< {end_ver}")

            if versions:
                result[product] = versions

    # affected가 없으면 configurations에서 추출
    if not result:
        for config in vuln.get('configurations', []):
            for node in config.get('nodes', []):
                for match in node.get('cpeMatch', []):
                    if match.get('vulnerable'):
                        cpe = match.get('criteria', '')
                        if cpe:
                            parts = cpe.split(':')
                            if len(parts) > 5:
                                product = parts[5] if parts[5] != '*' else 'Unknown'

                        start = match.get('versionStartIncluding') or match.get('versionStartExcluding', '')
                        end = match.get('versionEndIncluding') or match.get('versionEndExcluding', '')

                        if start or end:
                            if start and end:
                                version = f"{start} ~ {end}"
                            elif end:
                                version = f"<= {end}"
                            else:
                                version = f">= {start}"

                            if product not in result:
                                result[product] = []
                            result[product].append(version)

    return result

def compare_results(cve_id: str, excel_r: str, nvd_data: Dict):
    """파싱된 데이터와 원본 데이터 비교"""

    print(f"\n  NVD 원본 데이터:")
    nvd_affected = extract_nvd_affected(nvd_data)

    if nvd_affected:
        for product, versions in nvd_affected.items():
            for v in versions[:2]:  # 처음 2개만
                print(f"    {product} {v}")
    else:
        print(f"    (버전 정보 없음)")

    print(f"\n  파싱된 데이터 (R 열):")
    if excel_r:
        # 첫 2개만 출력
        parts = excel_r.split(';')[:2]
        for p in parts:
            print(f"    {p.strip()}")
        if len(excel_r.split(';')) > 2:
            print(f"    ... +{len(excel_r.split(';')) - 2}개")
    else:
        print(f"    (없음)")

    # 간단한 정확도 판정
    match = False
    if excel_r and nvd_affected:
        for product, versions in nvd_affected.items():
            for v in versions:
                if v.strip() in excel_r or product.lower() in excel_r.lower():
                    match = True
                    break

    return "✓ 일치" if match else "✗ 불일치"

# 메인 검증
print("=" * 100)
print("정확도 검증 - 원본 NVD 데이터 vs 파싱된 데이터")
print("=" * 100)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

# 검증할 행들 (CVE가 있는 행)
test_rows = [2, 7, 8, 9, 15, 20, 51]  # 다양한 행들

for row in test_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if not cve or not isinstance(cve, str) or not cve.startswith('CVE-'):
        continue

    print(f"\n\n{'─' * 100}")
    print(f"[행 {row}] {cve}")
    print(f"{'─' * 100}")

    # NVD 데이터 조회
    nvd_data = get_nvd_data(cve)

    if nvd_data:
        result = compare_results(cve, r_col, nvd_data)
        print(f"\n  검증 결과: {result}")
    else:
        print(f"  검증 불가 (NVD 조회 실패)")

print(f"\n\n{'=' * 100}")
print("검증 완료")
print("=" * 100)

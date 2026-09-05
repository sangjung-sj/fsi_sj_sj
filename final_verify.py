#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""최종 결과 확인"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_134651.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 110)
print("최종 결과 확인 - R 열에 제품명 포함")
print("=" * 110)

sample_rows = [2, 7, 8, 9, 15, 20]

for row in sample_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    print(f"\n[행 {row}] CVE: {cve}")
    if r_col:
        # R 열이 길면 줄바꿈
        if len(str(r_col)) > 80:
            print(f"  R 열: {str(r_col)[:80]}...")
        else:
            print(f"  R 열: {r_col}")
    else:
        print(f"  R 열: (없음)")
    print(f"  S 열: {s_col}")

print("\n" + "=" * 110)
print("✓ 완료. 제품명이 R 열에 포함되었습니다.")
print("=" * 110)

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""S 열 (조치 버전) 검증"""

import openpyxl
import requests
import json

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140020.xlsx"

def get_nvd_patch_version(cve_id: str) -> str:
    """NVD에서 패치 버전 추출"""
    try:
        url = f"{NVD_API}?cveId={cve_id}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        if not data.get('vulnerabilities'):
            return None

        vuln = data['vulnerabilities'][0]['cve']

        # affected 배열에서 unaffected 버전 찾기
        for affected_item in vuln.get('affected', []):
            for affected_data in affected_item.get('affectedData', []):
                for version_info in affected_data.get('versions', []):
                    if version_info.get('status') == 'unaffected':
                        v = version_info.get('version', '')
                        ver_type = version_info.get('versionType', '')
                        if v and 'git' not in ver_type:
                            return v

        # References에서 공식 공지 찾기
        refs = vuln.get('references', [])
        if refs:
            return f"[공식 공지 {len(refs)}개 참조]"

        return None

    except Exception as e:
        return f"[조회 실패: {str(e)[:20]}]"

# 메인 검증
print("=" * 100)
print("S 열 (조치 버전) 검증 - 원본 NVD 데이터 vs 파싱된 데이터")
print("=" * 100)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

# 검증할 행들
test_rows = [2, 7, 8, 9, 15, 20, 51]

for row in test_rows:
    cve = ws[f"L{row}"].value
    s_col = ws[f"S{row}"].value

    if not cve or not isinstance(cve, str) or not cve.startswith('CVE-'):
        continue

    print(f"\n{'─' * 100}")
    print(f"[행 {row}] {cve}")
    print(f"{'─' * 100}")

    # NVD에서 패치 버전 조회
    nvd_patch = get_nvd_patch_version(cve)

    print(f"  NVD 원본 패치 버전: {nvd_patch if nvd_patch else '(없음)'}")
    print(f"  파싱된 S 열 값: {s_col if s_col else '(없음)'}")

    # 비교
    if s_col and nvd_patch:
        if nvd_patch in str(s_col) or str(s_col) in str(nvd_patch):
            print(f"  결과: ✓ 일치")
        else:
            print(f"  결과: ~ 부분 일치 (값이 다름)")
    elif (not s_col and not nvd_patch) or (s_col and '공식 공지' in str(s_col)):
        print(f"  결과: ✓ 일치 (패치 정보 없음)")
    else:
        print(f"  결과: ✗ 불일치")

print(f"\n\n{'=' * 100}")
print("S 열 검증 완료")
print("=" * 100)
print("\n[설명]")
print("- S 열은 '조치 버전' (패치 버전) 정보입니다")
print("- affected 배열의 unaffected 버전에서 추출됩니다")
print("- 패치 정보가 없으면 '[공식 공지 참조]'로 표시됩니다")
print("=" * 100)

#!/usr/bin/env python3
import openpyxl
import requests

wb = openpyxl.load_workbook('넥스포즈 자료정리_updated_20260902_150402.xlsx')
ws = wb.active

NVD_API = 'https://services.nvd.nist.gov/rest/json/cves/2.0'

print('=' * 180)
print('[확인 필요] 표시된 행 분석')
print('=' * 180)

for row in range(2, ws.max_row + 1):
    r_col = ws[f'R{row}'].value
    s_col = ws[f'S{row}'].value
    cve = ws[f'L{row}'].value

    # S열이 "[확인 필요]"인 경우만
    if s_col and str(s_col).strip() == '[확인 필요]':
        print(f'\n[행 {row}] {cve}')
        print('-' * 180)

        if r_col:
            print(f'R열 (취약 버전): ✓ 있음')
            for line in str(r_col).split('\n')[:2]:
                print(f'  - {line}')
        else:
            print(f'R열 (취약 버전): ✗ 없음')

        print(f'S열 (패치 버전): ✓ [확인 필요]')

        # NVD에서 원본 데이터 확인
        if cve:
            try:
                url = f'{NVD_API}?cveId={cve}'
                resp = requests.get(url, timeout=10)
                data = resp.json()

                if data.get('vulnerabilities'):
                    vuln = data['vulnerabilities'][0]['cve']

                    # Affected Products 확인
                    affected_products = []
                    for affected_item in vuln.get('affected', []):
                        for affected_data in affected_item.get('affectedData', []):
                            product = affected_data.get('product', '')
                            unaffected_versions = []

                            for version_info in affected_data.get('versions', []):
                                if version_info.get('status') == 'unaffected':
                                    v = version_info.get('version', '')
                                    if v and v != '0':
                                        unaffected_versions.append(v)

                            if unaffected_versions:
                                affected_products.append({
                                    'product': product,
                                    'versions': unaffected_versions
                                })

                    print(f'\n📌 NVD에서 확인 가능한 정보:')
                    if affected_products:
                        print(f'Affected Products의 Unaffected 버전:')
                        for ap in affected_products:
                            print(f'  - {ap["product"]}: {", ".join(ap["versions"])}')
                    else:
                        print(f'Affected Products에 unaffected 버전 정보 없음')

                    print(f'\n🔗 확인 방법:')
                    print(f'  1. NVD 웹사이트: https://nvd.nist.gov/vuln/detail/{cve}')
                    print(f'  2. 페이지에서 "Affected Products" 섹션 → Unaffected 버전 확인')
                    print(f'  3. 또는 "Known Affected Software Configurations" 섹션에서')
                    print(f'     제품별 버전 범위를 보고 역으로 패치 버전 파악')

            except Exception as e:
                print(f'오류: {e}')

print(f'\n' + '=' * 180)
print('분석 결과:')
print('=' * 180)
print('''
[확인 필요]가 표시된 이유:
1. R열: Configurations에서 취약 버전 범위를 찾음
2. S열: Affected Products의 unaffected 버전이 없거나 제한적

해결 방법:
1️⃣ NVD 웹사이트 방문: https://nvd.nist.gov/vuln/detail/CVE-ID
2️⃣ "Affected Products" 섹션에서 Unaffected 버전 확인
3️⃣ 원본 문서나 보안 공지 확인
4️⃣ 제품 개발사 웹사이트에서 보안 권고 확인
''')

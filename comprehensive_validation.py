#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""종합 정확도 검증 - R 열과 S 열"""

import openpyxl
import requests
import json

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140710.xlsx"

def get_nvd_data(cve_id: str):
    """NVD에서 원본 데이터 조회"""
    try:
        url = f"{NVD_API}?cveId={cve_id}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"    ✗ NVD 조회 실패: {str(e)[:30]}")
        return None

def extract_nvd_info(data) -> dict:
    """NVD 원본 데이터에서 제품명, 취약 버전, 패치 버전 추출"""
    if not data or not data.get('vulnerabilities'):
        return {'product': None, 'affected': [], 'patch': None}

    vuln = data['vulnerabilities'][0]['cve']
    product = None
    affected_versions = []
    patch_versions = []

    # affected 배열에서 추출
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            product = affected_data.get('product', 'Unknown')

            for version_info in affected_data.get('versions', []):
                status = version_info.get('status', '')
                v = version_info.get('version', '')
                less_eq = version_info.get('lessThanOrEqual', '')
                less_than = version_info.get('lessThan', '')

                # 취약 버전
                if status == 'affected' and (less_eq or less_than):
                    end_ver = less_eq or less_than
                    if end_ver and '*' not in end_ver:
                        if v and v != '0':
                            affected_versions.append(f"{v} ~ {end_ver}")
                        else:
                            affected_versions.append(f"<= {end_ver}" if less_eq else f"< {end_ver}")

                # 패치 버전
                if status == 'unaffected' and v and v != '0':
                    ver_type = version_info.get('versionType', '')
                    if 'git' not in ver_type and not v.isdigit():
                        patch_versions.append(v)

    patch = patch_versions[0] if patch_versions else None

    return {
        'product': product,
        'affected': affected_versions,
        'patch': patch
    }

def validate_row(row_num, cve, r_col, s_col, nvd_data) -> dict:
    """행 검증"""
    result = {
        'row': row_num,
        'cve': cve,
        'r_check': False,
        's_check': False,
        'overall': False
    }

    if not nvd_data:
        result['status'] = '✗ NVD 조회 실패'
        return result

    nvd_info = extract_nvd_info(nvd_data)
    product = nvd_info['product']
    affected = nvd_info['affected']
    patch = nvd_info['patch']

    # R 열 검증
    r_check = False
    if r_col and product and affected:
        # 제품명 확인
        if product.lower() in str(r_col).lower():
            # 버전 정보 확인
            for v in affected[:2]:
                if v in str(r_col):
                    r_check = True
                    break
    result['r_check'] = r_check

    # S 열 검증
    s_check = False
    if s_col:
        if product and product.lower() in str(s_col).lower():
            if patch:
                s_check = patch in str(s_col)
            elif '공식 공지' in str(s_col):
                s_check = True
    result['s_check'] = s_check

    result['overall'] = r_check and s_check
    result['status'] = ('✓' if result['overall'] else '~') if (r_check or s_check) else '✗'

    # 상세 정보
    result['nvd'] = {
        'product': product,
        'affected_sample': affected[:2] if affected else [],
        'patch': patch
    }

    return result

# 메인 검증
print("=" * 140)
print("종합 정확도 검증 - R 열과 S 열 비교 분석")
print("=" * 140)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

# 검증할 행들 (CVE가 있는 행)
test_rows = [2, 7, 8, 9, 15, 20, 51]
results = []

print("\n[상세 검증 결과]\n")

for row in test_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if not cve or not isinstance(cve, str) or not cve.startswith('CVE-'):
        continue

    nvd_data = get_nvd_data(cve)
    validation = validate_row(row, cve, r_col, s_col, nvd_data)
    results.append(validation)

    print(f"{validation['status']} [행 {row}] {cve}")
    print(f"    R 열: {'✓' if validation['r_check'] else '✗'} {str(r_col)[:80]}{'...' if r_col and len(str(r_col)) > 80 else ''}")
    print(f"    S 열: {'✓' if validation['s_check'] else '✗'} {str(s_col)[:80]}{'...' if s_col and len(str(s_col)) > 80 else ''}")

    if nvd_data:
        nvd = validation['nvd']
        print(f"    NVD 제품: {nvd['product']}")
        if nvd['affected_sample']:
            print(f"    NVD 취약: {', '.join(nvd['affected_sample'][:1])}")
        if nvd['patch']:
            print(f"    NVD 패치: {nvd['patch']}")
    print()

# 요약 통계
print("=" * 140)
print("검증 요약")
print("=" * 140)

total = len(results)
overall_pass = sum(1 for r in results if r['overall'])
r_pass = sum(1 for r in results if r['r_check'])
s_pass = sum(1 for r in results if r['s_check'])

print(f"\n전체 행 수:         {total}")
print(f"R 열 정확도:        {r_pass}/{total} ({100*r_pass//total if total > 0 else 0}%)")
print(f"S 열 정확도:        {s_pass}/{total} ({100*s_pass//total if total > 0 else 0}%)")
print(f"전체 정확도:        {overall_pass}/{total} ({100*overall_pass//total if total > 0 else 0}%)")

print("\n" + "=" * 140)

if overall_pass == total:
    print("✓ 모든 행이 정확하게 파싱되었습니다!")
elif overall_pass >= total * 0.8:
    print("~ 대부분의 행이 정확하게 파싱되었습니다. 일부 개선 항목 있음.")
else:
    print(f"✗ 개선 필요: {total - overall_pass}개 행에서 검증 실패")

print("=" * 140)
print("\n[검증 설명]")
print("- R 열: 제품명 + 취약 버전이 NVD와 일치하는지 확인")
print("- S 열: 제품명 + 패치 버전이 NVD와 일치하는지 확인")
print("- ✓ 일치: R/S 열 모두 정확")
print("- ~ 부분일치: R 또는 S 열 중 하나만 정확")
print("- ✗ 불일치: 데이터가 정확하지 않음")
print("=" * 140)

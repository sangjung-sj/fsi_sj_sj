#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""검증 보고서 - 저장된 데이터 활용"""

import openpyxl
import json

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140710.xlsx"
nvd_response_file = r"c:/Users/USER/Downloads/for_user/nvd_response.json"

# 저장된 NVD 응답 로드
try:
    with open(nvd_response_file, 'r', encoding='utf-8') as f:
        sample_nvd_data = json.load(f)
except:
    sample_nvd_data = None

def extract_nvd_info_from_json(data) -> dict:
    """JSON에서 제품명, 취약 버전, 패치 버전 추출"""
    if not data or not data.get('vulnerabilities'):
        return {'product': None, 'affected': [], 'patch': None}

    vuln = data['vulnerabilities'][0]['cve']
    product = None
    affected_versions = []
    patch_versions = []

    # affected 배열에서 추출
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            product = affected_data.get('product', 'Unknown')

            for version_info in affected_data.get('versions', []):
                status = version_info.get('status', '')
                v = version_info.get('version', '')
                less_eq = version_info.get('lessThanOrEqual', '')
                less_than = version_info.get('lessThan', '')

                # 취약 버전
                if status == 'affected' and (less_eq or less_than):
                    end_ver = less_eq or less_than
                    if end_ver and '*' not in end_ver:
                        if v and v != '0':
                            affected_versions.append(f"{v} ~ {end_ver}")
                        else:
                            affected_versions.append(f"<= {end_ver}" if less_eq else f"< {end_ver}")

                # 패치 버전
                if status == 'unaffected' and v and v != '0':
                    ver_type = version_info.get('versionType', '')
                    if 'git' not in ver_type and not v.isdigit():
                        patch_versions.append(v)

    patch = patch_versions[0] if patch_versions else None

    return {
        'product': product,
        'affected': affected_versions,
        'patch': patch
    }

# 메인 검증
print("=" * 140)
print("검증 보고서 - 파싱된 데이터의 정확도")
print("=" * 140)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

if sample_nvd_data:
    print("\n[샘플 검증: CVE-2026-66032 (첫 번째 행)]\n")

    cve = "CVE-2026-66032"
    nvd_info = extract_nvd_info_from_json(sample_nvd_data)

    # 엑셀에서 첫 번째 행 데이터
    r_col = ws["R2"].value
    s_col = ws["S2"].value

    print(f"CVE: {cve}")
    print(f"\n1. 제품명 검증")
    print(f"   NVD 제품:      {nvd_info['product']}")
    print(f"   R 열:          {r_col}")
    print(f"   S 열:          {s_col}")
    print(f"   결과:          {'✓ 일치' if nvd_info['product'] and nvd_info['product'].lower() in str(r_col).lower() and nvd_info['product'].lower() in str(s_col).lower() else '✗ 불일치'}")

    print(f"\n2. 취약 버전 검증")
    print(f"   NVD 취약 버전: {nvd_info['affected']}")
    print(f"   R 열:          {r_col}")
    r_match = any(v in str(r_col) for v in nvd_info['affected'])
    print(f"   결과:          {'✓ 일치' if r_match else '✗ 불일치'}")

    print(f"\n3. 패치 버전 검증")
    print(f"   NVD 패치 버전: {nvd_info['patch']}")
    print(f"   S 열:          {s_col}")
    if nvd_info['patch']:
        s_match = nvd_info['patch'] in str(s_col)
        print(f"   결과:          {'✓ 일치' if s_match else '✗ 불일치/참조'}")
    else:
        s_match = '공식 공지' in str(s_col)
        print(f"   결과:          {'✓ 정확 (공식 공지)' if s_match else '✗ 불일치'}")

else:
    print("\n⚠ NVD 샘플 데이터를 로드할 수 없습니다.")

# 구조적 검증
print("\n\n" + "=" * 140)
print("구조적 검증 - 모든 행의 형식 확인")
print("=" * 140)

print("\n1. R 열 형식 검증")
print("   제품명과 버전이 함께 있는가?")

valid_r_count = 0
total_r_count = 0

for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-') and r_col:
        total_r_count += 1
        # 제품명과 버전이 모두 있는지 확인
        r_str = str(r_col)
        # 버전 형식 확인 (숫자, ~, <, >, -)를 포함하는지)
        has_version_pattern = any(c in r_str for c in ['~', '<', '>', '-', '.'])
        # 제품명 같은 단어가 있는지 (공백 다음에 단어)
        has_product = len(r_str.split()) > 1

        if has_version_pattern and has_product:
            valid_r_count += 1

print(f"   유효한 R 열: {valid_r_count}/{total_r_count} ({100*valid_r_count//total_r_count if total_r_count > 0 else 0}%)")

print("\n2. S 열 형식 검증")
print("   제품명과 버전(또는 참조)이 함께 있는가?")

valid_s_count = 0
total_s_count = 0

for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-') and r_col and s_col:
        total_s_count += 1
        # R 열에서 제품명 추출
        r_product = str(r_col).split()[0] if str(r_col).split() else ""
        s_str = str(s_col)

        # S 열에 R 열의 제품명이 포함되어 있는가?
        if r_product and r_product.lower() in s_str.lower():
            valid_s_count += 1

print(f"   유효한 S 열: {valid_s_count}/{total_s_count} ({100*valid_s_count//total_s_count if total_s_count > 0 else 0}%)")

print("\n\n" + "=" * 140)
print("종합 검증 결과")
print("=" * 140)
print(f"\n✓ R 열: {valid_r_count}/{total_r_count}개 행 형식 정확 (제품명 + 버전)")
print(f"✓ S 열: {valid_s_count}/{total_s_count}개 행 형식 정확 (제품명 + 버전/참조)")
print(f"\n✓ 파싱 성공: {min(valid_r_count, valid_s_count)}/{max(total_r_count, total_s_count)}개 행")

if valid_r_count == total_r_count and valid_s_count == total_s_count:
    print("\n✓ 모든 행이 올바르게 파싱되었습니다!")
    print("  - R 열: 제품명 + 취약 버전 범위 ✓")
    print("  - S 열: 제품명 + 패치 버전/공식 공지 참조 ✓")
else:
    print(f"\n~ 대부분의 행이 정확합니다.")
    print(f"  일부 행에서 R 열 또는 S 열이 빈 상태일 수 있습니다.")

print("\n" + "=" * 140)

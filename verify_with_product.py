#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""제품명 포함 결과 확인"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_134326.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 100)
print("제품명 포함된 결과 확인")
print("=" * 100)

sample_rows = [2, 7, 8, 9, 15, 20]

for row in sample_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    print(f"\n[행 {row}] CVE: {cve}")
    print(f"  R 열 (취약 버전): {r_col}")
    print(f"  S 열 (조치 버전): {s_col}")

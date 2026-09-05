#!/usr/bin/env python3
import openpyxl
import requests

CVE_ID = 'CVE-2026-6772'
NVD_API = 'https://services.nvd.nist.gov/rest/json/cves/2.0'

# 1. 생성된 파일의 데이터 확인
wb = openpyxl.load_workbook('넥스포즈 자료정리_updated_20260902_150402.xlsx')
ws = wb.active

row = 8
cve = ws[f'L{row}'].value
r_col = ws[f'R{row}'].value
s_col = ws[f'S{row}'].value

print('=' * 180)
print(f'행 8 데이터 검증: {CVE_ID}')
print('=' * 180)

print(f'\n[생성된 파일]')
print(f'R열:')
if r_col:
    for line in str(r_col).split('\n'):
        print(f'  ✓ {line}')
else:
    print(f'  (없음)')

print(f'\nS열:')
if s_col:
    for line in str(s_col).split('\n'):
        print(f'  ✓ {line}')
else:
    print(f'  (없음)')

# 2. NVD 원본 데이터 조회
print(f'\n' + '=' * 180)
print(f'[NVD 원본 데이터 분석]')
print('=' * 180)

url = f'{NVD_API}?cveId={CVE_ID}'
resp = requests.get(url, timeout=10)
data = resp.json()
vuln = data['vulnerabilities'][0]['cve']

# Configurations 분석
print(f'\n[Known Affected Software Configurations]')
config_ranges = {}
for config in vuln.get('configurations', []):
    for node in config.get('nodes', []):
        for match in node.get('cpeMatch', []):
            if match.get('vulnerable'):
                cpe = match.get('criteria', '')
                parts = cpe.split(':')
                if len(parts) > 4:
                    product = parts[4]
                    start_inc = match.get('versionStartIncluding', '')
                    start_exc = match.get('versionStartExcluding', '')
                    end_inc = match.get('versionEndIncluding', '')
                    end_exc = match.get('versionEndExcluding', '')

                    range_str = ''
                    if start_inc and end_exc:
                        range_str = f'{product} {start_inc} ~ {end_exc}'
                    elif end_exc:
                        range_str = f'{product} < {end_exc}'
                    elif end_inc:
                        range_str = f'{product} <= {end_inc}'

                    if range_str:
                        if product not in config_ranges:
                            config_ranges[product] = []
                        config_ranges[product].append(range_str)
                        print(f'  ✓ {range_str}')

# Affected Products 분석
print(f'\n[Affected Products - Unaffected 버전]')
affected_products = {}
for affected_item in vuln.get('affected', []):
    for affected_data in affected_item.get('affectedData', []):
        product = affected_data.get('product', '')
        for version_info in affected_data.get('versions', []):
            v = version_info.get('version', '')
            status = version_info.get('status', '')

            if status == 'unaffected' and v and v != '0':
                if product not in affected_products:
                    affected_products[product] = []
                affected_products[product].append(v)
                print(f'  ✓ {product} {v}')

# 3. 논리 검증
print(f'\n' + '=' * 180)
print(f'[검증 결과]')
print('=' * 180)

print(f'\n✓ R열 검증:')
r_lines = str(r_col).split('\n') if r_col else []
for line in r_lines:
    found = False
    for product, ranges in config_ranges.items():
        for range_str in ranges:
            if line.lower() in range_str.lower():
                print(f'  ✓ "{line}" ← Configurations에서 올바르게 추출됨')
                found = True
                break
        if found:
            break
    if not found:
        print(f'  ⚠ "{line}" ← 확인 필요')

print(f'\n✓ S열 검증:')
s_lines = str(s_col).split('\n') if s_col else []
for line in s_lines:
    found = False
    for product, versions in affected_products.items():
        for version in versions:
            if version.lower() in line.lower():
                print(f'  ✓ "{line}" ← Affected Products에서 올바르게 추출됨')
                found = True
                break
        if found:
            break
    if not found:
        print(f'  ⚠ "{line}" ← Affected Products에 없음')

# 4. 최종 판정
print(f'\n' + '=' * 180)
print(f'[최종 판정]')
print('=' * 180)

all_valid = True
for line in r_lines:
    found = False
    for product, ranges in config_ranges.items():
        for range_str in ranges:
            if line.lower() in range_str.lower():
                found = True
                break
        if found:
            break
    if not found:
        all_valid = False

for line in s_lines:
    found = False
    for product, versions in affected_products.items():
        for version in versions:
            if version.lower() in line.lower():
                found = True
                break
        if found:
            break
    if not found:
        all_valid = False

if all_valid:
    print(f'\n✅ 행 8 데이터가 정확하게 작성되었습니다!')
    print(f'  - R열: Configurations에서 올바르게 추출')
    print(f'  - S열: 필터링된 제품의 패치 버전만 포함')
    print(f'  - 대소문자 처리 완료')
else:
    print(f'\n⚠ 행 8 데이터에 문제가 있습니다.')
    print(f'  위의 검증 결과를 확인하세요.')

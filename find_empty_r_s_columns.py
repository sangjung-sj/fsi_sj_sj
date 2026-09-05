#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""R, S열이 공란인 행 찾기"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_142748.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 150)
print("R, S열이 공란인 행 분석")
print("=" * 150)

# 1. CVE가 있는데 R열이 공란인 행
print("\n[1️⃣ CVE가 있는데 R열이 공란]")
print("-" * 150)

cve_but_no_r = []
for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-') and not r_col:
        cve_but_no_r.append((row, cve))

if cve_but_no_r:
    print(f"\n발견: {len(cve_but_no_r)}개\n")
    for row, cve in cve_but_no_r:
        print(f"  [행 {row}] {cve}")
else:
    print(f"\n✓ 없음 (모든 CVE의 R열이 채워짐)")

# 2. CVE가 있는데 S열이 공란인 행
print("\n\n[2️⃣ CVE가 있는데 S열이 공란]")
print("-" * 150)

cve_but_no_s = []
for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    s_col = ws[f"S{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-') and not s_col:
        cve_but_no_s.append((row, cve))

if cve_but_no_s:
    print(f"\n발견: {len(cve_but_no_s)}개\n")
    for row, cve in cve_but_no_s:
        print(f"  [행 {row}] {cve}")
else:
    print(f"\n✓ 없음 (모든 CVE의 S열이 채워짐)")

# 3. CVE가 없는데 Project가 있는 행 (수동 확인 필요)
print("\n\n[3️⃣ CVE가 없는데 Project(D열)가 있는 행 - GitHub 검색 대상]")
print("-" * 150)

no_cve_but_project = []
for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    ghsa = ws[f"M{row}"].value
    ant = ws[f"N{row}"].value
    project = ws[f"D{row}"].value
    r_col = ws[f"R{row}"].value

    # CVE, GHSA, ANT가 모두 없는데 Project가 있는 경우
    if (not cve or (isinstance(cve, str) and not cve.startswith('CVE-'))) and \
       (not ghsa or (isinstance(ghsa, str) and not ghsa.startswith('GHSA-'))) and \
       (not ant or (isinstance(ant, str) and not ant.startswith('ANT-'))) and \
       project:
        no_cve_but_project.append({
            'row': row,
            'project': project,
            'r_filled': bool(r_col),
            'r_value': r_col
        })

if no_cve_but_project:
    print(f"\n발견: {len(no_cve_but_project)}개\n")
    for item in no_cve_but_project:
        print(f"  [행 {item['row']}] {item['project']}")
        r_status = '✓ 채워짐' if item['r_filled'] else '❌ 공란'
        r_preview = f"({str(item['r_value'])[:50]}...)" if item['r_filled'] else ""
        print(f"    R열: {r_status} {r_preview}")
else:
    print(f"\n✓ 없음")

# 4. 아예 아무것도 없는 행
print("\n\n[4️⃣ 완전히 공란인 행 (수동 확인)]")
print("-" * 150)

completely_empty = []
for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    ghsa = ws[f"M{row}"].value
    ant = ws[f"N{row}"].value
    project = ws[f"D{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    # 모든 열이 공란
    if not any([cve, ghsa, ant, project, r_col, s_col]):
        completely_empty.append(row)

if completely_empty:
    print(f"\n발견: {len(completely_empty)}개\n")
    for row in completely_empty[:5]:  # 처음 5개만
        print(f"  [행 {row}] 완전 공란")
    if len(completely_empty) > 5:
        print(f"  ... 등 {len(completely_empty) - 5}개 더")
else:
    print(f"\n✓ 없음")

# 종합 통계
print("\n\n" + "=" * 150)
print("종합 통계")
print("=" * 150)

total_rows = ws.max_row - 1  # 헤더 제외
cve_count = 0
r_filled = 0
s_filled = 0

for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-'):
        cve_count += 1
        if r_col:
            r_filled += 1
        if s_col:
            s_filled += 1

print(f"""
전체 데이터 행:     {total_rows}개
├─ CVE 행:         {cve_count}개
├─ R열 채워짐:     {r_filled}/{cve_count} ({100*r_filled//cve_count if cve_count > 0 else 0}%)
├─ S열 채워짐:     {s_filled}/{cve_count} ({100*s_filled//cve_count if cve_count > 0 else 0}%)
└─ R&S 모두:       {min(r_filled, s_filled)}/{cve_count} ({100*min(r_filled, s_filled)//cve_count if cve_count > 0 else 0}%)

⚠️ 공란 현황:
├─ CVE 있는데 R만 공란: {len(cve_but_no_r)}개
├─ CVE 있는데 S만 공란: {len(cve_but_no_s)}개
└─ CVE 없는데 Project 있음: {len(no_cve_but_project)}개
""")

print("=" * 150)

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""최종 정확도 검증"""

import openpyxl
import requests

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140516.xlsx"

def get_nvd_data(cve_id: str):
    try:
        url = f"{NVD_API}?cveId={cve_id}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except:
        return None

def extract_patch_version(data) -> str:
    """NVD에서 패치 버전 추출"""
    if not data or not data.get('vulnerabilities'):
        return None

    vuln = data['vulnerabilities'][0]['cve']

    # affected 배열에서 unaffected 버전 찾기
    valid_versions = []
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            for version_info in affected_data.get('versions', []):
                if version_info.get('status') == 'unaffected':
                    v = version_info.get('version', '')
                    ver_type = version_info.get('versionType', '')
                    # git commit이나 전체 숫자(timestamp)는 제외
                    if v and 'git' not in ver_type and not v.isdigit():
                        valid_versions.append(v)

    return valid_versions[0] if valid_versions else None

# 메인
print("=" * 110)
print("최종 정확도 검증 - R 열과 S 열")
print("=" * 110)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

test_rows = [2, 7, 8, 9, 15, 20, 51]
matches = 0
total = 0

for row in test_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if not cve or not isinstance(cve, str) or not cve.startswith('CVE-'):
        continue

    total += 1

    nvd_data = get_nvd_data(cve)
    nvd_patch = extract_patch_version(nvd_data) if nvd_data else None

    # R 열 검증
    r_match = False
    if nvd_data and r_col and nvd_data.get('vulnerabilities'):
        vuln = nvd_data['vulnerabilities'][0]['cve']
        affected = vuln.get('affected', [])
        if affected:
            for item in affected:
                for data in item.get('affectedData', []):
                    product = data.get('product', '')
                    if product and product.lower() in str(r_col).lower():
                        r_match = True
                        break

    # S 열 검증
    s_match = False
    if nvd_patch and s_col:
        if nvd_patch == s_col:
            s_match = True
    elif not nvd_patch and s_col and '공식 공지' in str(s_col):
        s_match = True

    row_match = r_match and s_match
    if row_match:
        matches += 1

    result = "✓" if row_match else "~"
    print(f"\n[{result} 행 {row}] {cve}")
    print(f"  R 열: {str(r_col)[:70]}..." if r_col and len(str(r_col)) > 70 else f"  R 열: {r_col}")
    print(f"  S 열: {s_col}")
    if nvd_patch:
        print(f"  NVD 패치: {nvd_patch}")
        print(f"  S 열 일치: {'✓' if s_match else '✗'}")

print(f"\n\n{'=' * 110}")
print(f"검증 결과: {matches}/{total} 행 검증 완료 ({100*matches//total}% 정확도)")
print("=" * 110)

if matches == total:
    print("✓ 모든 행이 정확합니다!")
else:
    print(f"~ {total - matches}개 행에서 개선 필요")

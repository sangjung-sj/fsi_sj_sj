#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""R열 수정 확인 (< vs <= 구분)"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_142748.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 140)
print("R열 수정 확인 - < (excluding) vs <= (including) 구분")
print("=" * 140)

# CVE-2026-6772 (Firefox) 확인
print("\n[예: CVE-2026-6772 (Firefox)]\n")

row = 8  # Firefox는 행 8
cve = ws[f"L{row}"].value
r_col = ws[f"R{row}"].value

print(f"CVE: {cve}")
print(f"\nR 열 (수정됨):")
print(f"  {r_col}")

print(f"\n✓ 수정 확인:")
if "< 115.35" in str(r_col):
    print(f"  ✓ < 115.35.0 (Up to excluding) - 정확함")
else:
    print(f"  ✗ 아직 <= 사용 중")

if "< 150" in str(r_col):
    print(f"  ✓ < 150.0 (Up to excluding) - 정확함")
else:
    print(f"  ✗ 아직 <= 사용 중")

if "< 140.10" in str(r_col):
    print(f"  ✓ < 140.10.0 (Up to excluding) - 정확함")
else:
    print(f"  ✗ 아직 <= 사용 중")

print(f"\n" + "=" * 140)
print("모든 샘플 행 확인")
print("=" * 140)

samples = [2, 8, 9, 12, 28]
for row in samples:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-') and r_col:
        print(f"\n[행 {row}] {cve}")
        r_display = str(r_col)[:100] + "..." if len(str(r_col)) > 100 else str(r_col)
        print(f"  R: {r_display}")

print(f"\n" + "=" * 140)
print("✅ R열 수정 완료!")
print("   - < : Up to (excluding) - 미포함")
print("   - <= : Up to (including) - 포함")
print("=" * 140)

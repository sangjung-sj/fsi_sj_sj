#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""최종 결과 - 모든 행 요약"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_142231.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 150)
print("최종 완성 - 모든 행 매핑 결과")
print("=" * 150)

# 데이터 통계
total_rows = ws.max_row - 1  # 헤더 제외
cve_count = 0
r_filled = 0
s_filled = 0

for row in range(2, ws.max_row + 1):
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-'):
        cve_count += 1
        if r_col:
            r_filled += 1
        if s_col:
            s_filled += 1

print(f"\n📊 통계:")
print(f"  전체 데이터 행:     {total_rows}개")
print(f"  CVE 정보:          {cve_count}개")
print(f"  R 열 (취약 버전):   {r_filled}개 ({100*r_filled//cve_count if cve_count > 0 else 0}%)")
print(f"  S 열 (unaffected):  {s_filled}개 ({100*s_filled//cve_count if cve_count > 0 else 0}%)")

print(f"\n📝 샘플 데이터 (처음 5개 CVE):\n")

sample_count = 0
for row in range(2, ws.max_row + 1):
    if sample_count >= 5:
        break

    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    if cve and isinstance(cve, str) and cve.startswith('CVE-'):
        sample_count += 1
        print(f"[행 {row}] {cve}")
        if r_col:
            r_display = str(r_col)[:90] + "..." if len(str(r_col)) > 90 else str(r_col)
            print(f"  R: {r_display}")
        else:
            print(f"  R: (없음)")

        if s_col:
            s_display = str(s_col)[:90] + "..." if len(str(s_col)) > 90 else str(s_col)
            print(f"  S: {s_display}")
        else:
            print(f"  S: (없음)")
        print()

print("=" * 150)
print(f"✅ 완료!")
print(f"파일: 넥스포즈 자료정리_updated_20260902_142231.xlsx")
print("=" * 150)
print("\n형식:")
print("  R 열: 제품명 + 취약 버전 범위")
print("  S 열: 제품명 + unaffected 버전들 또는 [공식 공지 참조]")
print("=" * 150)

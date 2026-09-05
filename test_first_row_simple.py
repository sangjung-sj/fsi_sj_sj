#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""첫 번째 행만 테스트"""

import openpyxl
import requests
import json
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

excel_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리.xlsx"
NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"

# 엑셀 파일 로드
print("=" * 80)
print("엑셀 파일 로드 중...")
print("=" * 80)
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 첫 번째 데이터 행 읽기
row_idx = 2
cve = ws[f"L{row_idx}"].value
ghsa = ws[f"M{row_idx}"].value
ant = ws[f"N{row_idx}"].value
project = ws[f"D{row_idx}"].value

print(f"\n첫 번째 데이터 행 (행 {row_idx}):")
print(f"  Project (D): {project}")
print(f"  CVE (L): {cve}")
print(f"  GHSA (M): {ghsa}")
print(f"  ANT (N): {ant}")

# CVE 정보 조회 테스트
if cve and isinstance(cve, str) and cve.startswith("CVE-"):
    print(f"\n{'=' * 80}")
    print(f"NVD API 테스트: {cve}")
    print("=" * 80)

    try:
        url = f"{NVD_API}?cveId={cve}"
        print(f"URL: {url}")

        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        print(f"✓ API 응답 받음 (Status: {resp.status_code})")
        print(f"  Vulnerabilities 개수: {len(data.get('vulnerabilities', []))}")

        # JSON 전체 저장
        output_file = r"c:/Users/USER/Downloads/for_user/nvd_response.json"
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"  ✓ 전체 응답 저장: {output_file}")

        if data.get('vulnerabilities'):
            vuln = data['vulnerabilities'][0]['cve']

            print(f"\nCVE 기본 정보:")
            print(f"  ID: {vuln.get('id')}")
            desc = vuln.get('descriptions', [{}])[0].get('value', 'N/A')
            print(f"  Description: {desc[:100]}...")

            print(f"\nConfigurations 분석:")
            configs = vuln.get('configurations', [])
            print(f"  총 configs: {len(configs)}")

            affected_list = []

            for i, config in enumerate(configs[:1]):  # 첫 번째만
                print(f"\n  Config {i}:")
                nodes = config.get('nodes', [])
                print(f"    Nodes: {len(nodes)}")

                for j, node in enumerate(nodes[:1]):  # 첫 번째만
                    print(f"    Node {j}:")
                    cpe_matches = node.get('cpeMatch', [])
                    print(f"      CPE Matches: {len(cpe_matches)}")

                    for k, match in enumerate(cpe_matches[:2]):  # 처음 2개만
                        print(f"\n      Match {k}:")
                        print(f"        vulnerable: {match.get('vulnerable')}")
                        print(f"        criteria: {match.get('criteria')[:80]}...")

                        if match.get('vulnerable'):
                            vr = match.get('versionRange', {})
                            if vr:
                                print(f"        versionRange: {vr}")
                                start = vr.get('versionStartIncluding') or vr.get('versionStartExcluding')
                                end = vr.get('versionEndExcluding') or vr.get('versionEndIncluding')
                                print(f"          → Start: {start}, End: {end}")
                                if start or end:
                                    version_info = f"{start}~{end}" if (start and end) else (start or end)
                                    affected_list.append(version_info)
                                    print(f"          ✓ 추출됨: {version_info}")
                            else:
                                print(f"        versionRange: 없음")
                                # CPE에서 추출
                                cpe = match.get('criteria', '')
                                parts = cpe.split(':')
                                print(f"        CPE parts: {parts[:6]}")
                                if len(parts) > 5:
                                    version = parts[5]
                                    if version and version != '*':
                                        affected_list.append(version)
                                        print(f"          ✓ CPE에서 추출: {version}")

            print(f"\n최종 결과:")
            affected_str = "; ".join(set(affected_list)) if affected_list else None
            print(f"  Affected versions: {affected_str}")

        else:
            print("✗ Vulnerabilities 없음")

    except requests.exceptions.RequestException as e:
        print(f"✗ HTTP 오류: {e}")
    except Exception as e:
        print(f"✗ 오류: {e}")
        import traceback
        traceback.print_exc()

elif ghsa and isinstance(ghsa, str) and ghsa.startswith("GHSA-"):
    print(f"\n[GitHub Advisory 확인됨] {ghsa}")

elif project:
    print(f"\n[Project 확인됨] {project}")

else:
    print("\n⚠ CVE, GHSA, Project, ANT 정보 모두 없음")

print("\n" + "=" * 80)
print("테스트 완료")
print("=" * 80)

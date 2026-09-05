#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""행 8 NVD 원본 데이터 vs 파싱된 데이터 상세 비교"""

import requests
import json

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
CVE_ID = "CVE-2026-6772"

print("=" * 160)
print(f"NVD 원본 데이터 상세 분석: {CVE_ID}")
print("=" * 160)

# NVD에서 데이터 조회
url = f"{NVD_API}?cveId={CVE_ID}"
try:
    resp = requests.get(url, timeout=10)
    data = resp.json()

    if not data.get('vulnerabilities'):
        print("❌ CVE 정보 없음")
        exit()

    vuln = data['vulnerabilities'][0]['cve']

    # 1. Affected 배열 분석
    print("\n[1️⃣ AFFECTED 배열 - 영향받는 제품/버전]")
    print("=" * 160)

    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            product = affected_data.get('product', 'Unknown')
            print(f"\n제품: {product}")

            print(f"\n버전 정보:")
            for version_info in affected_data.get('versions', []):
                v = version_info.get('version', '')
                status = version_info.get('status', '')
                less_eq = version_info.get('lessThanOrEqual', '')
                less_than = version_info.get('lessThan', '')
                ver_type = version_info.get('versionType', '')

                print(f"  version={v} | status={status} | lessThanOrEqual={less_eq} | lessThan={less_than} | type={ver_type}")

    # 2. Configuration 분석
    print("\n\n[2️⃣ CONFIGURATION - 취약한 버전 범위]")
    print("=" * 160)

    for config in vuln.get('configurations', []):
        for node in config.get('nodes', []):
            for match in node.get('cpeMatch', []):
                if match.get('vulnerable'):
                    cpe = match.get('criteria', '')
                    start_inc = match.get('versionStartIncluding', '')
                    start_exc = match.get('versionStartExcluding', '')
                    end_inc = match.get('versionEndIncluding', '')
                    end_exc = match.get('versionEndExcluding', '')

                    print(f"\nCPE: {cpe}")
                    print(f"  versionStartIncluding: {start_inc if start_inc else '(없음)'}")
                    print(f"  versionStartExcluding: {start_exc if start_exc else '(없음)'}")
                    print(f"  versionEndIncluding:   {end_inc if end_inc else '(없음)'}")
                    print(f"  versionEndExcluding:   {end_exc if end_exc else '(없음)'}")

                    # 논리 정리
                    if start_inc and end_exc:
                        print(f"  → {start_inc} <= 버전 < {end_exc}")
                    elif end_exc:
                        print(f"  → 버전 < {end_exc} (미포함)")
                    elif end_inc:
                        print(f"  → 버전 <= {end_inc} (포함)")

    # 3. 논리적 분석
    print("\n\n[3️⃣ 논리적 정리]")
    print("=" * 160)

    print("\n✅ 취약한 버전 (R열에 들어가야 함):")

    affected_versions = []
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            product = affected_data.get('product', '')
            for version_info in affected_data.get('versions', []):
                if version_info.get('status') == 'affected':
                    v = version_info.get('version', '')
                    less_eq = version_info.get('lessThanOrEqual', '')
                    less_than = version_info.get('lessThan', '')

                    if less_eq and '*' not in less_eq:
                        if v and v != '0':
                            affected_versions.append(f"{product} {v} ~ {less_eq}")
                        else:
                            affected_versions.append(f"{product} <= {less_eq}")
                    elif less_than and '*' not in less_than:
                        if v and v != '0':
                            affected_versions.append(f"{product} {v} ~ {less_than}")
                        else:
                            affected_versions.append(f"{product} < {less_than}")

    for av in affected_versions:
        print(f"  - {av}")

    print("\n✅ 안전한 버전 (S열에 들어가야 함):")

    safe_versions = []
    for affected_item in vuln.get('affected', []):
        for affected_data in affected_item.get('affectedData', []):
            product = affected_data.get('product', '')
            for version_info in affected_data.get('versions', []):
                if version_info.get('status') == 'unaffected':
                    v = version_info.get('version', '')
                    ver_type = version_info.get('versionType', '')

                    print(f"  - {product} {v} (type: {ver_type})")

                    # 포함 조건
                    if v and 'git' not in ver_type:
                        # 정수만 있는 경우 (예: 150)
                        if v.isdigit():
                            safe_versions.append(f"{product} {v}")
                        # 일반 버전 (예: 115.35, 140.10)
                        else:
                            safe_versions.append(f"{product} {v}")

    print("\n🔍 S열에 넣을 안전한 버전:")
    for sv in safe_versions:
        print(f"  - {sv}")

    print("\n\n[4️⃣ 현재 파싱 vs 정확한 데이터]")
    print("=" * 160)

    import openpyxl
    output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_142748.xlsx"
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    r_col = ws["R8"].value
    s_col = ws["S8"].value

    print(f"\n현재 파싱된 데이터:")
    print(f"  R열: {r_col}")
    print(f"  S열: {s_col}")

    print(f"\n정확해야 할 데이터:")
    print(f"  R열: {'; '.join(affected_versions)}")
    print(f"  S열: {'; '.join(safe_versions)}")

    print(f"\n문제점:")
    if "150" not in (s_col or ""):
        print(f"  ❌ S열에 Firefox 150이 없음! (있어야 함)")
    else:
        print(f"  ✓ S열에 Firefox 150이 있음")

    print("\n" + "=" * 160)

except Exception as e:
    print(f"❌ 오류: {e}")

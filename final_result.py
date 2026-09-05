#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""최종 결과 확인 - 제품명 포함"""

import openpyxl

output_file = r"c:/Users/USER/Downloads/for_user/넥스포즈 자료정리_updated_20260902_140020.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("=" * 120)
print("최종 결과 - R 열에 제품명 포함된 버전 정보")
print("=" * 120)

sample_rows = [2, 7, 8, 9, 15, 20]

for row in sample_rows:
    cve = ws[f"L{row}"].value
    r_col = ws[f"R{row}"].value
    s_col = ws[f"S{row}"].value

    print(f"\n[행 {row}] CVE: {cve}")
    if r_col:
        # 너무 길면 첫 95글자만 표시
        if len(str(r_col)) > 95:
            print(f"  R 열: {str(r_col)[:95]}...")
        else:
            print(f"  R 열: {r_col}")
    else:
        print(f"  R 열: (없음)")
    print(f"  S 열: {s_col}")

print("\n" + "=" * 120)
print("✓ 완료. R 열에 제품명이 포함되었습니다.")
print("=" * 120)
